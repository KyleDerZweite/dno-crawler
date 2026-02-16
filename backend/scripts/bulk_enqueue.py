#!/usr/bin/env python3
"""
Bulk-enqueue crawl jobs for DNOs listed in an XLSX or CSV file.

Reads DNO names from the file, resolves them against the API,
and enqueues crawl jobs for each matched DNO via HTTP.

Requires a DNO_API_KEY environment variable with a valid API key (dno_ prefix).

Usage:
    # From backend/ directory:
    export DNO_API_KEY=dno_...
    uv run python scripts/bulk_enqueue.py ../data/VNB.xlsx --year 2025
    uv run python scripts/bulk_enqueue.py ../data/VNB.xlsx --year 2025 --job-type extract
    uv run python scripts/bulk_enqueue.py ../data/VNB.xlsx --year 2025 --data-type netzentgelte
    uv run python scripts/bulk_enqueue.py ../data/VNB.xlsx --year 2025 --dry-run
"""

import argparse
import csv
import io
import os
import re
import sys
from difflib import SequenceMatcher
from pathlib import Path

import httpx

# Strip German legal form suffixes so "Foo GmbH" vs "Foo AG" aren't penalised.
_LEGAL_SUFFIX_RE = re.compile(
    r"\s*\b(gmbh|mbh|ag|kg|co\.\s*kg|gmbh\s*&\s*co\.\s*kg|se|e\.?\s*v\.?|ohg|ug)\s*$",
    re.IGNORECASE,
)
# Strip year suffixes like "(2025)" that appear in spreadsheet names.
_YEAR_SUFFIX_RE = re.compile(r"\s*\(\d{4}\)\s*$")
# Collapse punctuation, hyphens, dots, and extra whitespace.
_PUNCT_RE = re.compile(r"[\s\-\.\/\+,]+")

# Similarity thresholds (after normalisation).
# Short names (< 15 chars) need a higher bar because a single-character
# difference (SWB vs SWO, ED vs EVB) is a different company.
FUZZY_THRESHOLD_SHORT = 0.9  # for normalised names < 15 chars
FUZZY_THRESHOLD_LONG = 0.8  # for normalised names >= 15 chars


def _normalize_name(name: str) -> str:
    """Normalise a DNO name for similarity comparison."""
    n = _LEGAL_SUFFIX_RE.sub("", name).strip()
    n = _YEAR_SUFFIX_RE.sub("", n).strip()
    n = _PUNCT_RE.sub(" ", n).strip().lower()
    return n


def _is_fuzzy_match(a: str, b: str) -> tuple[bool, float]:
    """Check if two DNO names are a plausible match after normalisation.

    Returns (accepted, score).  A match is accepted when:
    - SequenceMatcher ratio >= threshold (tiered: stricter for short names), OR
    - The shorter normalised name is a prefix of the longer one (>= 3 chars),
      handling parent → subsidiary abbreviations like "E.ON" → "E.ON Energie
      Deutschland".
    """
    na = _normalize_name(a)
    nb = _normalize_name(b)
    ratio = SequenceMatcher(None, na, nb).ratio()
    short, long = (na, nb) if len(na) <= len(nb) else (nb, na)
    threshold = FUZZY_THRESHOLD_SHORT if len(short) < 15 else FUZZY_THRESHOLD_LONG
    if ratio >= threshold:
        return True, ratio
    # Prefix check: shorter name starts the longer name (parent -> subsidiary)
    if long.startswith(short) and len(short) >= 3:
        return True, ratio
    return False, ratio


def parse_xlsx(file_path: Path) -> list[str]:
    """Extract DNO names from XLSX. Uses first column or a known header."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []

    # Detect header column
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    col_idx = 0
    for candidate in ("verteilnetzbetreiber", "dno_name", "name", "dno"):
        if candidate in header:
            col_idx = header.index(candidate)
            break

    names: list[str] = []
    for row in rows[1:]:
        if col_idx < len(row) and row[col_idx]:
            val = str(row[col_idx]).strip()
            if val:
                names.append(val)

    wb.close()
    return names


def parse_csv(file_path: Path) -> list[str]:
    """Extract DNO names from CSV. Uses first column or a known header."""
    text = file_path.read_text(encoding="utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    fieldnames = [f.strip().lower() for f in (reader.fieldnames or [])]

    col_name: str | None = None
    for candidate in ("verteilnetzbetreiber", "dno_name", "name", "dno"):
        if candidate in fieldnames:
            col_name = candidate
            break

    names: list[str] = []
    for row in reader:
        lower_row = {k.strip().lower(): v for k, v in row.items()}
        if col_name:
            val = lower_row.get(col_name, "").strip()
        else:
            val = next(iter(row.values()), "").strip()
        if val:
            names.append(val)

    return names


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Bulk-enqueue crawl jobs for DNOs from an XLSX/CSV file"
    )
    parser.add_argument(
        "file",
        type=Path,
        help="Path to XLSX or CSV file with DNO names",
    )
    parser.add_argument(
        "--year",
        type=int,
        required=True,
        help="Target year for crawl jobs",
    )
    parser.add_argument(
        "--data-type",
        choices=["all", "netzentgelte", "hlzf"],
        default="all",
        help="Data type to crawl (default: all)",
    )
    parser.add_argument(
        "--job-type",
        choices=["full", "crawl", "extract"],
        default="full",
        help="Job type (default: full)",
    )
    parser.add_argument(
        "--priority",
        type=int,
        default=5,
        choices=range(1, 11),
        metavar="1-10",
        help="Job priority 1-10 (default: 5)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Show what would be enqueued without creating jobs",
    )
    parser.add_argument(
        "--api-url",
        default="http://localhost:8000/api/v1",
        help="API base URL (default: http://localhost:8000/api/v1)",
    )

    args = parser.parse_args()

    api_key = os.environ.get("DNO_API_KEY", "")
    if not api_key:
        print("Error: DNO_API_KEY environment variable is required")
        print("Create an API key via the Admin Dashboard, then:")
        print("  export DNO_API_KEY=dno_...")
        return 1

    if not args.file.exists():
        print(f"Error: File not found: {args.file}")
        return 1

    # Parse file
    ext = args.file.suffix.lower()
    if ext == ".xlsx":
        dno_names = parse_xlsx(args.file)
    elif ext == ".csv":
        dno_names = parse_csv(args.file)
    else:
        print(f"Error: Unsupported file type '{ext}'. Use .xlsx or .csv")
        return 1

    # Deduplicate preserving order
    dno_names = list(dict.fromkeys(dno_names))
    print(f"Parsed {len(dno_names)} unique DNO names from {args.file.name}")

    if not dno_names:
        print("No DNO names found in file")
        return 1

    # Set up HTTP client
    headers = {"Authorization": f"Bearer {api_key}"}
    client = httpx.Client(base_url=args.api_url, headers=headers, timeout=30.0)

    # Verify API key works
    try:
        me_resp = client.get("/auth/me")
        if me_resp.status_code != 200:
            print(f"Error: API key authentication failed (HTTP {me_resp.status_code})")
            client.close()
            return 1
        me_data = me_resp.json()
        print(f"Authenticated as: {me_data.get('data', {}).get('name', 'unknown')}")
    except httpx.ConnectError:
        print(f"Error: Could not connect to API at {args.api_url}")
        client.close()
        return 1

    # Fetch all DNOs from the API
    print("Fetching DNO list from API...")
    all_dnos: list[dict] = []
    page = 1
    while True:
        resp = client.get("/dnos/", params={"page": page, "per_page": 200})
        if resp.status_code != 200:
            print(f"Error: Failed to fetch DNOs (HTTP {resp.status_code})")
            client.close()
            return 1
        data = resp.json()
        dnos_page = data.get("data", [])
        if not dnos_page:
            break
        all_dnos.extend(dnos_page)
        meta = data.get("meta", {})
        if page >= meta.get("total_pages", 1):
            break
        page += 1

    print(f"Fetched {len(all_dnos)} DNOs from API")

    # Build lookup: lowercase name -> (slug, name)
    name_to_dno: dict[str, tuple[str, str]] = {}
    for dno in all_dnos:
        name_to_dno[dno["name"].lower()] = (dno["slug"], dno["name"])

    # Phase 1: Exact case-insensitive matching
    matched: dict[str, str] = {}  # input_name -> slug
    unmatched: list[str] = []
    for name in dno_names:
        hit = name_to_dno.get(name.lower())
        if hit:
            matched[name] = hit[0]
        else:
            unmatched.append(name)

    # Phase 2: Fuzzy search via API for unmatched names (same trigram search as the UI)
    # Results are filtered by client-side similarity to reject false positives.
    fuzzy_matched: dict[str, tuple[str, str, float]] = {}  # input -> (slug, db_name, score)
    fuzzy_rejected: dict[str, tuple[str, float]] = {}  # input -> (db_name, score)
    not_found: list[str] = []
    if unmatched:
        print(f"Exact match: {len(matched)}, trying fuzzy search for {len(unmatched)} remaining...")
        for name in unmatched:
            resp = client.get("/dnos/", params={"q": name, "per_page": 1})
            if resp.status_code == 200:
                results = resp.json().get("data", [])
                if results:
                    top = results[0]
                    accepted, score = _is_fuzzy_match(name, top["name"])
                    if accepted:
                        fuzzy_matched[name] = (top["slug"], top["name"], score)
                        matched[name] = top["slug"]
                    else:
                        fuzzy_rejected[name] = (top["name"], score)
                        not_found.append(name)
                else:
                    not_found.append(name)
            else:
                not_found.append(name)

    if fuzzy_matched:
        print(f"Fuzzy matched {len(fuzzy_matched)} additional DNOs:")
        for input_name, (slug, db_name, score) in fuzzy_matched.items():
            print(f"  {input_name} -> {db_name} ({slug}) [{score:.0%}]")

    if fuzzy_rejected:
        print(f"Rejected {len(fuzzy_rejected)} low-confidence fuzzy matches:")
        for input_name, (db_name, score) in fuzzy_rejected.items():
            print(f"  {input_name} ~> {db_name} [{score:.0%}]")

    print(f"\nTotal matched: {len(matched)}, Not found: {len(not_found)}")
    if not_found:
        for nf in not_found:
            print(f"  Not found: {nf}")

    if not matched:
        print("No matching DNOs found. Exiting.")
        client.close()
        return 1

    if args.dry_run:
        print(
            f"\n[DRY RUN] Would create {len(matched)} {args.job_type} jobs "
            f"(year={args.year}, data_type={args.data_type}, priority={args.priority})"
        )
        for name, slug in matched.items():
            print(f"  {slug}: {name}")
        client.close()
        return 0

    # Enqueue crawl jobs via API
    created = 0
    skipped = 0
    failed = 0

    for _input_name, slug in matched.items():
        payload = {
            "year": args.year,
            "data_type": args.data_type,
            "job_type": args.job_type,
            "priority": args.priority,
        }
        try:
            resp = client.post(f"/dnos/{slug}/crawl", json=payload)
            if resp.status_code == 200:
                job_data = resp.json()
                job_id = job_data.get("data", {}).get("job_id", "?")
                print(f"  Created: {slug} -> job {job_id}")
                created += 1
            elif resp.status_code == 409:
                print(f"  Skipped: {slug} (already has pending/running job)")
                skipped += 1
            elif resp.status_code == 404:
                print(f"  Not found: {slug}")
                failed += 1
            else:
                detail = resp.json().get("detail", resp.text[:100])
                print(f"  Failed: {slug} (HTTP {resp.status_code}: {detail})")
                failed += 1
        except httpx.HTTPError as e:
            print(f"  Error: {slug} ({e})")
            failed += 1

    client.close()

    print(
        f"\nDone: {created} created, {skipped} skipped, {failed} failed, {len(not_found)} not found"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
